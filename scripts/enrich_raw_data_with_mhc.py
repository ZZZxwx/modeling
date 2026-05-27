from __future__ import annotations

import argparse
import hashlib
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
PYTHON_ROOT = REPO_ROOT / "python"
if str(PYTHON_ROOT) not in sys.path:
    sys.path.insert(0, str(PYTHON_ROOT))
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

try:
    from scripts.spec_excel_summary_common import (
        percent_as_ratio,
        sheet_key_values,
        value_as_float,
    )
except ModuleNotFoundError:
    from spec_excel_summary_common import (
        percent_as_ratio,
        sheet_key_values,
        value_as_float,
    )


MHC_COLUMNS = [
    "mhc_time_ms",
    "mhc_step_share",
    "mhc_delta_step_ms",
    "mhc_delta_step_pct",
    "mhc_report_path",
    "mhc_status",
    "mhc_error",
]

NO_MHC_COLUMNS = [
    "no_mhc_match_status",
    "no_mhc_step_time_ms",
    "no_mhc_tp",
    "no_mhc_cp",
    "no_mhc_pp",
    "no_mhc_ep",
    "no_mhc_dp",
    "no_mhc_pp_schedule",
    "no_mhc_vpp_chunks",
    "no_mhc_cp_kind",
    "no_mhc_micro_batch",
    "no_mhc_global_batch",
    "no_mhc_zero_stage",
    "no_mhc_optimizer",
    "no_mhc_source_name",
    "no_mhc_source_file",
]

CONFIG_DEFAULTS: dict[str, object] = {
    "hw": "nvidia_h100_sxm",
    "seq_len": 4096,
    "world_size": 1,
    "tp": 1,
    "cp": 1,
    "pp": 1,
    "ep": 1,
    "dp": 1,
    "micro_batch": 1,
    "global_batch": 0,
    "pp_schedule": "1f1b",
    "vpp_chunks": 1,
    "cp_kind": "none",
    "zero_stage": 0,
    "recompute": "none",
    "optimizer": "adam",
    "tp_overlap": "none",
    "ep_overlap": False,
    "dualbatch": False,
    "quant_preset": None,
}

INT_KEYS = {
    "world_size", "seq_len", "tp", "cp", "pp", "ep", "dp",
    "micro_batch", "global_batch", "vpp_chunks", "zero_stage",
}
BOOL_KEYS = {"ep_overlap", "dualbatch"}
FINGERPRINT_KEYS = [
    "model", "hw", "world_size", "seq_len", "tp", "cp", "pp", "ep", "dp",
    "micro_batch", "global_batch", "pp_schedule", "vpp_chunks", "cp_kind",
    "zero_stage", "recompute", "optimizer", "tp_overlap", "ep_overlap",
    "dualbatch", "quant_preset",
]


@dataclass
class MHCBackfillResult:
    mhc_time_ms: float = 0.0
    mhc_step_share: float = 0.0
    mhc_delta_step_ms: float = 0.0
    mhc_delta_step_pct: float = 0.0
    report_path: str = ""
    status: str = "ok"
    error: str = ""


Runner = Callable[[dict[str, object], Path], MHCBackfillResult]


def enrich_workbook(
    *,
    input_path: str | Path,
    output_path: str | Path,
    report_dir: str | Path,
    default_model: str = "deepseek_v4_pro",
    no_mhc_summary_path: str | Path | None = None,
    existing_report_dir: str | Path | None = None,
    runner: Runner | None = None,
) -> Path:
    input_path = Path(input_path)
    output_path = Path(output_path)
    report_dir = Path(report_dir)
    existing_report_dir = Path(existing_report_dir) if existing_report_dir else None
    runner = runner or run_spec_report_for_config
    existing_reports_by_name = (
        build_existing_report_name_index(existing_report_dir)
        if existing_report_dir is not None else {}
    )

    wb = openpyxl.load_workbook(input_path)
    if "raw_data" not in wb.sheetnames:
        raise ValueError(f"{input_path} does not contain a raw_data sheet")
    ws = wb["raw_data"]

    headers = _header_map(ws)
    _ensure_columns(ws, headers, MHC_COLUMNS)
    if no_mhc_summary_path is not None:
        _ensure_columns(ws, headers, NO_MHC_COLUMNS)
    headers = _header_map(ws)
    no_mhc_by_hw = load_no_mhc_summary(no_mhc_summary_path) if no_mhc_summary_path else {}

    cache: dict[str, MHCBackfillResult] = {}
    for row_idx in range(2, ws.max_row + 1):
        config = row_config(ws, headers, row_idx, default_model)
        key = config_fingerprint(config)
        report_path = (
            existing_report_dir / f"{key}.xlsx"
            if existing_report_dir is not None
            else report_dir / f"{key}.xlsx"
        )
        if existing_report_dir is not None and not report_path.exists():
            named_report = find_existing_report_by_row_name(config, existing_reports_by_name)
            if named_report is not None:
                report_path = named_report
        if key not in cache:
            cache[key] = _load_or_run_mhc_result(
                config=config,
                report_path=report_path,
                existing_report_dir=existing_report_dir,
                runner=runner,
            )
        _write_result(ws, headers, row_idx, cache[key])
        if no_mhc_summary_path is not None:
            _write_no_mhc_result(ws, headers, row_idx, config, no_mhc_by_hw)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _load_or_run_mhc_result(
    *,
    config: dict[str, object],
    report_path: Path,
    existing_report_dir: Path | None,
    runner: Runner,
) -> MHCBackfillResult:
    if existing_report_dir is not None:
        if not report_path.exists():
            return MHCBackfillResult(
                report_path=str(report_path),
                status="missing_report",
                error=f"existing report not found: {report_path}",
            )
        try:
            return extract_mhc_result(report_path)
        except Exception as exc:
            return MHCBackfillResult(
                report_path=str(report_path),
                status="error",
                error=str(exc),
            )
    try:
        return runner(config, report_path)
    except Exception as exc:
        return MHCBackfillResult(
            report_path=str(report_path),
            status="error",
            error=str(exc),
        )


def row_config(ws, headers: dict[str, int], row_idx: int,
               default_model: str) -> dict[str, object]:
    config = dict(CONFIG_DEFAULTS)
    config["model"] = _cell_value(ws, headers, row_idx, "model") or default_model
    for optional_key in ("name", "config_name", "project", "file"):
        raw_optional = _cell_value(ws, headers, row_idx, optional_key)
        if raw_optional is not None and raw_optional != "":
            config[optional_key] = raw_optional
    for key in CONFIG_DEFAULTS:
        raw = _cell_value(ws, headers, row_idx, key)
        if raw is None or raw == "":
            continue
        config[key] = _coerce_config_value(key, raw)
    return config


def config_fingerprint(config: dict[str, object]) -> str:
    canonical = {key: config.get(key) for key in FINGERPRINT_KEYS}
    payload = json.dumps(canonical, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()[:16]


def build_existing_report_name_index(report_dir: Path) -> dict[str, Path | None]:
    index: dict[str, Path | None] = {}
    for path in sorted(report_dir.rglob("*.xls*"), key=lambda p: str(p).lower()):
        if path.name.startswith("~$"):
            continue
        key = normalize_report_match_name(path.stem)
        if not key:
            continue
        if key in index:
            index[key] = None
        else:
            index[key] = path
    return index


def find_existing_report_by_row_name(
    config: dict[str, object],
    existing_reports_by_name: dict[str, Path | None],
) -> Path | None:
    for field in ("name", "config_name", "project", "file"):
        raw = config.get(field)
        if raw is None or raw == "":
            continue
        key = normalize_report_match_name(str(raw))
        match = existing_reports_by_name.get(key)
        if match is not None:
            return match
    return None


def normalize_report_match_name(name: str) -> str:
    stem = Path(str(name)).stem.lower().strip()
    if stem.startswith("deepseek_v4_pro_"):
        stem = stem[len("deepseek_v4_pro_"):]
    if stem.endswith("_best"):
        stem = stem[:-len("_best")]
    parts = [p for p in re_split_underscores(stem) if p]
    if not parts:
        return ""
    parts[-1] = normalize_batch_token(parts[-1])
    return "_".join(parts)


def re_split_underscores(value: str) -> list[str]:
    return [p.strip() for p in value.replace("-", "_").split("_")]


def normalize_batch_token(token: str) -> str:
    text = token.strip().lower()
    if text in {"1", "1k", "1024"}:
        return "1024"
    if text.endswith("k"):
        try:
            return str(int(float(text[:-1]) * 1024))
        except ValueError:
            return text
    return text


def load_no_mhc_summary(path: str | Path | None) -> dict[str, dict[str, object]]:
    if path is None:
        return {}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Summary" not in wb.sheetnames:
        raise ValueError(f"{path} does not contain a Summary sheet")
    ws = wb["Summary"]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(v).strip() if v is not None else "" for v in next(rows)]
    except StopIteration:
        return {}
    index = {name: idx for idx, name in enumerate(header) if name}
    if "hw" not in index:
        raise ValueError(f"{path} Summary sheet does not contain an hw column")

    by_hw: dict[str, dict[str, object]] = {}
    duplicate: set[str] = set()
    for row in rows:
        if not row or all(v is None for v in row):
            continue
        hw = str(row[index["hw"]]).strip() if len(row) > index["hw"] and row[index["hw"]] is not None else ""
        if not hw:
            continue
        data = {
            col: row[idx] if idx < len(row) else None
            for col, idx in index.items()
        }
        if hw in by_hw:
            duplicate.add(hw)
        by_hw[hw] = data
    for hw in duplicate:
        by_hw[hw]["__duplicate__"] = True
    return by_hw


def run_spec_report_for_config(config: dict[str, object], report_path: Path) -> MHCBackfillResult:
    from zrt.hardware.registry import load as load_hw
    from zrt.training.io.excel_exporter import export_estimate_excel
    from zrt.training.ir.builders import build_graph
    from zrt.training.models.flops import op_cost
    from zrt.training.search.estimator import estimate
    from zrt.training.search.training_search_util import (
        _ceil_nodes_for_world_size,
        _inferred_gpus_per_node,
        _load_model_spec,
        _make_strategy_from_config,
        _system_from_hw,
    )

    model = _load_model_spec(str(config["model"]), quant_preset=config.get("quant_preset") or None)
    model.seq_len = int(config.get("seq_len", model.seq_len))

    hw = load_hw(str(config.get("hw", "nvidia_h100_sxm")))
    world_size = int(config.get("world_size", 1))
    gpus_per_node = _inferred_gpus_per_node(hw)
    system = _system_from_hw(
        hw,
        nodes=_ceil_nodes_for_world_size(world_size, gpus_per_node),
        gpus_per_node=gpus_per_node,
        world_size_override=world_size,
        host_mem_gb=float(config.get("host_mem_gb", 256.0) or 256.0),
    )

    strategy = _make_strategy_from_config(config)
    graph = build_graph(model, strategy)
    op_costs = {op.name: op_cost(op, model, system) for op in graph.ops}
    report = estimate(model, system, strategy, graph=graph)

    report_path.parent.mkdir(parents=True, exist_ok=True)
    export_estimate_excel(
        report=report,
        graph=graph,
        model=model,
        system=system,
        strategy=strategy,
        op_costs=op_costs,
        output_path=report_path,
    )
    return extract_mhc_result(report_path)


def extract_mhc_result(report_path: str | Path) -> MHCBackfillResult:
    report_path = Path(report_path)
    wb = openpyxl.load_workbook(report_path, data_only=True, read_only=True)
    mhc = sheet_key_values(wb, "MHC Analysis")
    cf = _counterfactual_values(wb)
    return MHCBackfillResult(
        mhc_time_ms=value_as_float(mhc.get("Total MHC Time")),
        mhc_step_share=percent_as_ratio(mhc.get("MHC Share of Step")),
        mhc_delta_step_ms=value_as_float(cf.get("Delta")),
        mhc_delta_step_pct=percent_as_ratio(cf.get("Delta %")) * 100.0,
        report_path=str(report_path),
        status="ok",
        error="",
    )


def _counterfactual_values(wb) -> dict[str, object]:
    if "MHC Analysis" not in wb.sheetnames:
        return {}
    ws = wb["MHC Analysis"]
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "Step Time (ms)":
            return {
                "HC On": row[1] if len(row) > 1 else None,
                "HC Off": row[2] if len(row) > 2 else None,
                "Delta": row[3] if len(row) > 3 else None,
                "Delta %": row[4] if len(row) > 4 else None,
            }
    return {}


def _header_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): idx
        for idx, cell in enumerate(ws[1], start=1)
        if cell.value is not None and str(cell.value).strip()
    }


def _ensure_columns(ws, headers: dict[str, int], columns: list[str]) -> None:
    next_col = ws.max_column + 1
    for col in columns:
        if col in headers:
            continue
        ws.cell(row=1, column=next_col, value=col)
        headers[col] = next_col
        next_col += 1


def _cell_value(ws, headers: dict[str, int], row_idx: int, key: str) -> object:
    col = headers.get(key)
    return None if col is None else ws.cell(row=row_idx, column=col).value


def _coerce_config_value(key: str, value: object) -> object:
    if key in INT_KEYS:
        return int(value)
    if key in BOOL_KEYS:
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes", "y"}
    return value


def _write_result(ws, headers: dict[str, int], row_idx: int,
                  result: MHCBackfillResult) -> None:
    values = {
        "mhc_time_ms": result.mhc_time_ms,
        "mhc_step_share": result.mhc_step_share,
        "mhc_delta_step_ms": result.mhc_delta_step_ms,
        "mhc_delta_step_pct": result.mhc_delta_step_pct,
        "mhc_report_path": result.report_path,
        "mhc_status": result.status,
        "mhc_error": result.error,
    }
    for key, value in values.items():
        ws.cell(row=row_idx, column=headers[key], value=value)


def _write_no_mhc_result(
    ws,
    headers: dict[str, int],
    row_idx: int,
    config: dict[str, object],
    no_mhc_by_hw: dict[str, dict[str, object]],
) -> None:
    hw = str(config.get("hw", "")).strip()
    row = no_mhc_by_hw.get(hw)
    if row is None:
        values = {"no_mhc_match_status": "missing"}
    elif row.get("__duplicate__"):
        values = {"no_mhc_match_status": "duplicate"}
    else:
        values = {
            "no_mhc_match_status": "ok",
            "no_mhc_step_time_ms": row.get("step_time_ms"),
            "no_mhc_tp": row.get("tp"),
            "no_mhc_cp": row.get("cp"),
            "no_mhc_pp": row.get("pp"),
            "no_mhc_ep": row.get("ep"),
            "no_mhc_dp": row.get("dp"),
            "no_mhc_pp_schedule": row.get("pp_schedule"),
            "no_mhc_vpp_chunks": row.get("vpp_chunks"),
            "no_mhc_cp_kind": row.get("cp_kind"),
            "no_mhc_micro_batch": row.get("micro_batch"),
            "no_mhc_global_batch": row.get("global_batch"),
            "no_mhc_zero_stage": row.get("zero_stage"),
            "no_mhc_optimizer": row.get("optimizer"),
            "no_mhc_source_name": row.get("name"),
            "no_mhc_source_file": row.get("file"),
        }
    for key in NO_MHC_COLUMNS:
        ws.cell(row=row_idx, column=headers[key], value=values.get(key))


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Backfill MHC metrics into the raw_data sheet of a search summary workbook."
    )
    parser.add_argument("--workbook", required=True, help="Input workbook with raw_data sheet.")
    parser.add_argument("--output", required=True, help="Output workbook path.")
    parser.add_argument("--report-dir", required=True, help="Directory for generated spec Excel reports.")
    parser.add_argument("--default-model", default="deepseek_v4_pro")
    parser.add_argument("--no-mhc-summary", help="Optional no-MHC summary .xlsx keyed by hw.")
    parser.add_argument(
        "--existing-report-dir",
        help="Optional directory of pre-generated reports named by config fingerprint. "
             "When set, reports are read instead of regenerated.",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    out = enrich_workbook(
        input_path=args.workbook,
        output_path=args.output,
        report_dir=args.report_dir,
        default_model=args.default_model,
        no_mhc_summary_path=args.no_mhc_summary,
        existing_report_dir=args.existing_report_dir,
    )
    print(f"Wrote enriched workbook to {out}")


if __name__ == "__main__":
    main()
