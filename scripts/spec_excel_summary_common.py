from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

import openpyxl


EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
MHC_KINDS = {"hc_expand", "mhc_pre", "mhc_post", "mhc_head"}


def clean_report_name(path_or_name: str | Path) -> str:
    stem = Path(path_or_name).stem
    if stem.startswith("deepseek_v4_pro_"):
        stem = stem[len("deepseek_v4_pro_"):]
    if stem.endswith("_best"):
        stem = stem[:-len("_best")]
    return stem


def normalize_report_match_name(path_or_name: str | Path) -> str:
    stem = clean_report_name(path_or_name).lower().strip()
    parts = [p.strip() for p in stem.replace("-", "_").split("_") if p.strip()]
    if not parts:
        return ""
    parts[-1] = normalize_batch_token(parts[-1])
    return "_".join(parts)


def normalize_batch_token(token: str) -> str:
    text = token.strip().lower()
    if text in {"1", "1k", "1024"}:
        return "1024"
    if text.endswith("k"):
        try:
            return str(int(float(text[:-1]) * 1024))
        except ValueError:
            return text
    return text


def find_excel_files(input_dir: str | Path) -> list[Path]:
    root = Path(input_dir)
    files = [
        p for p in root.rglob("*")
        if p.is_file() and p.suffix.lower() in EXCEL_SUFFIXES and not p.name.startswith("~$")
    ]
    return sorted(files, key=lambda p: str(p).lower())


def load_workbook(path: str | Path):
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def sheet_key_values(wb, sheet_name: str) -> dict[str, object]:
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    out: dict[str, object] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        if not key:
            continue
        out[key] = row[1] if len(row) > 1 else None
    return out


def value_as_float(value: object, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.upper() == "N/A":
        return default
    text = text.replace(",", "")
    match = re.search(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", text)
    return float(match.group(0)) if match else default


def percent_as_ratio(value: object, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        val = float(value)
        return val / 100.0 if val > 1.0 else val
    text = str(value).strip()
    val = value_as_float(text, default)
    return val / 100.0 if "%" in text else val


def write_summary_xlsx(rows: list[dict[str, object]], columns: Iterable[str],
                       output: str | Path) -> Path:
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    cols = list(columns)
    ws.append(cols)
    for row in rows:
        ws.append([row.get(col) for col in cols])

    for idx, col in enumerate(cols, 1):
        max_len = len(col)
        for cell in ws.iter_cols(min_col=idx, max_col=idx, values_only=True):
            for value in cell:
                max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = min(max_len + 2, 80)

    wb.save(output_path)
    return output_path


def step_time_ms(wb) -> float:
    summary = sheet_key_values(wb, "Summary")
    return value_as_float(summary.get("Step Time"))


def strategy_values(wb) -> dict[str, object]:
    raw = sheet_key_values(wb, "Strategy")
    mapping = {
        "TP": "tp",
        "CP": "cp",
        "PP": "pp",
        "EP": "ep",
        "DP": "dp",
        "PP Schedule": "pp_schedule",
        "VPP Chunks": "vpp_chunks",
        "CP Kind": "cp_kind",
        "Micro Batch": "micro_batch",
        "Global Batch": "global_batch",
        "Zero Stage": "zero_stage",
        "Optimizer": "optimizer",
    }
    return {out_key: raw.get(in_key) for in_key, out_key in mapping.items()}


def mhc_time_from_ops(wb) -> float:
    if "Ops" not in wb.sheetnames:
        return 0.0
    ws = wb["Ops"]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(v).strip() if v is not None else "" for v in next(rows)]
    except StopIteration:
        return 0.0

    kind_idx = _find_header(header, "Kind")
    latency_idx = _find_header(header, "Latency")
    if kind_idx is None or latency_idx is None:
        return 0.0

    latency_us = 0.0
    for row in rows:
        if len(row) <= max(kind_idx, latency_idx):
            continue
        kind = str(row[kind_idx]).strip() if row[kind_idx] is not None else ""
        if kind in MHC_KINDS:
            latency_us += value_as_float(row[latency_idx])
    return latency_us / 1000.0


def _find_header(header: list[str], target: str) -> int | None:
    target_l = target.lower()
    for idx, name in enumerate(header):
        if target_l in name.lower():
            return idx
    return None
